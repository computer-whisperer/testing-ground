import sys
import os
import shutil
import openpyxl
import re
import urllib.parse
import subprocess
import time
from os import path


def main():
    machine_list = sys.argv[1]
    curr_build_dir = sys.argv[2]
    curr_dir_path, curr_dir_name = path.split(curr_build_dir)
    new_machine_folder = sys.argv[3]
    new_machine_list = path.join(new_machine_folder, "machine list.xlsx")
    new_machine_list_dir, _ = path.split(new_machine_list)
    new_build_dir = path.join(new_machine_folder, "builds")
    new_misc_dir = path.join(new_machine_folder, "unmatched build sheets")
    for dir in (new_machine_folder, new_build_dir, new_misc_dir):
        if not path.exists(dir):
            os.mkdir(dir)

    sorted_objects = []
    files_transferred = 0



    wb = openpyxl.load_workbook(machine_list)
    ws = wb.active

    company_col = ""
    customer_headers = ["company", "customer", "built for"]
    for cell in ws["1"]:
        if cell.value is None:
            continue
        for h in customer_headers:
            if cell.value.lower() == h:
                company_col = cell.column
                break
        else:
            continue
        break
    else:
        raise ValueError("could not find company column")

    for cell in ws["A"]:
        if cell.hyperlink is not None:
            link = urllib.parse.unquote(cell.hyperlink.target)
            dir_base, filename = path.split(link)
            sorted_objects.append(filename)
            dir_pre_base, dir_name = path.split(dir_base)
            if dir_name == curr_dir_name:
                curr_build_sheet_path = path.join(curr_build_dir, filename)
            else:
                print("ehh, problemo....")
                print(link)
                print(dir_name)
                print(curr_dir_name)
                continue

            # Create a new directory for the build
            serial_num = cell.value
            end_user = ws[company_col][cell.row-1].value
            if not isinstance(end_user, str):
                new_build_sheet_folder = serial_num
            else:
                new_build_sheet_folder = "{} {}".format(serial_num, re.sub('[^A-Za-z0-9-\s]+', '', end_user)).strip()
            new_build_sheet_dir = path.join(new_build_dir, new_build_sheet_folder)
            if not os.path.exists(new_build_sheet_dir):
                os.mkdir(new_build_sheet_dir)

            # If the linked object is a directory, copy all contents over to the new directory
            if path.isdir(curr_build_sheet_path):
                if len(os.listdir(curr_build_sheet_path)) > 0:
                    for filename in os.listdir(curr_build_sheet_path):
                        old_file = path.join(curr_build_sheet_path, filename)
                        new_file = path.join(new_build_sheet_dir, filename)
                        if not path.exists(new_file):
                            shutil.copy(old_file, new_file)
                            files_transferred += 1
                else:
                    curr_build_sheet_path = path.join(curr_build_sheet_path, "missing_file")

            # Otherwise try to copy the linked build sheet to the new directory
            if path.isfile(curr_build_sheet_path):
                new_build_sheet_path = path.join(new_build_sheet_dir, filename)
                if not os.path.exists(new_build_sheet_path):
                    if os.path.exists(curr_build_sheet_path):
                        shutil.copy(curr_build_sheet_path, new_build_sheet_path)
                        files_transferred += 1
                    else:
                        # If we have problems finding it, let the user fix the problem
                        if len(os.listdir(new_build_sheet_dir)) == 0:
                            print("\n\nWe seem to have a problem. I was trying to copy {} to {}, but {} doesn't seem to exist!".format(
                                curr_build_sheet_path, new_build_sheet_path, curr_build_sheet_path))
                            print("Please place the correct build sheet in {}".format(new_build_sheet_dir))
                            subprocess.Popen(r'explorer "{}"'.format(new_build_sheet_dir))

                            while len(os.listdir(new_build_sheet_dir)) == 0:
                                input("Hit enter when you have found the correct build sheet")
                            new_build_sheet_path = os.path.join(new_build_sheet_dir, os.listdir(new_build_sheet_dir)[0])
                            print("Found {}".format(new_build_sheet_path))

            # Resolve relative path and save to machine list
            rel_new_build_sheet_dir = os.path.relpath(new_build_sheet_dir, new_machine_list_dir)
            cell.hyperlink.target = rel_new_build_sheet_dir
    print("Finished transferring {} files into the new tree, saving xlsx".format(files_transferred))
    wb.save(new_machine_list)

    print("Scanning for leftovers")
    leftover_ct = 0
    # Scan for files in curr_build_dir that we did not copy over and save in another directory
    for filename in os.listdir(curr_build_dir):
        if filename not in sorted_objects and not filename.startswith("~"):
            old_file = path.join(curr_build_dir, filename)
            new_file = path.join(new_misc_dir, filename)
            if not path.exists(new_file):
                leftover_ct += 1
                if path.isfile(old_file):
                    shutil.copy(old_file, new_file)
                else:
                    shutil.copytree(old_file, new_file)
    print("Found {} leftovers, put them in {}".format(leftover_ct, new_misc_dir))

if __name__ == "__main__":
    main()
